import csv
import pandas as pd
import openpyxl
import os
from datetime import *
from dotenv import *
load_dotenv()
import pandas as pd
import re
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import shutil

config = os.environ

runid = datetime.today().strftime("%Y%m%d")

def shrink_df(df, type):
    if type == 'benefits':
        selected_columns = ['benefit_name', 'variant_id', 'carrier_id', 'plan_variant_name', 'hios_plan_id', 'SERFF Coinsurance', 'SERFF Copay', 'SERFF Copay Value', 'Rule Value', 'match_result', 'Result Description']
    else:
        selected_columns = ['benefit_name', 'variant_id', 'carrier_id', 'plan_variant_name', 'cost_sharing_variant', f'SERFF {type}', f'SERFF {type} Value', 'Rule Value', 'match_result', 'Result Description']
    df = df[selected_columns].copy()
    return df


def gen_excel(benefits, duct, moop):
    path = config['deliverable_path']
    rule_path = config['rule.template']
    file_path = path + f'nm_serff_comp_{runid}.xlsx' 
    # create copy of OG rule template
    shutil.copy2(rule_path, file_path)

    book = load_workbook(file_path)

    # save and close workbook
    book.save(file_path)
    book.close()

    # distribute dataframes to corresponding sheets
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        benefits.to_excel(writer, sheet_name='Benefits', index=False)
        duct.to_excel(writer, sheet_name='Deductible', index=False)
        moop.to_excel(writer, sheet_name='MOOP', index=False)

    book = load_workbook(file_path)

    # save and close workbook
    book.save(file_path)
    book.close()

    for sheet_name in book.sheetnames[1:]:
        recolor_sheets(sheet_name)
    add_autofilter_to_sheets(file_path)
        
    print(f'Results have been saved to: {file_path}')


def add_autofilter_to_sheets(file_path):
    book = load_workbook(file_path)
    for sheet_name in book.sheetnames[1:]:
        worksheet = book[sheet_name]
        worksheet.auto_filter.ref = worksheet.dimensions
    book.save(file_path)
    print(f'Autofilters have been added to all sheets in the file: {file_path}')


def recolor_sheets(sheet_name):
    path = config['deliverable_path']
    file_path = path + f'nm_serff_comp_{runid}.xlsx'

    # connect to file
    workbook = openpyxl.load_workbook(file_path)

    # assign sheet
    worksheet = workbook[sheet_name]

    # create/define colors
    light_green_fill = PatternFill(start_color='00CCFFCC', end_color='00CCFFCC', fill_type='solid')
    light_red_fill = PatternFill(start_color='00FFCCCC', end_color='00FFCCCC', fill_type='solid')
    light_blue_fill = PatternFill(start_color='0099CCFF', end_color='0099CCFF', fill_type='solid')
    light_orange_fill = PatternFill(start_color='00FFCC99', end_color='00FFCC99', fill_type='solid')

    # get headers
    headers = [cell.value for cell in worksheet[1]]

    # iterate over the columns and format the cells based on the column headers
    for col_idx, header in enumerate(headers, 1):
        if header == 'match_result':
            # iterate over the cells in the 'match_result' column
            for cell in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                if cell[0].value is True:
                    cell[0].fill = light_green_fill
                elif cell[0].value is False:
                    cell[0].fill = light_red_fill

        elif header == 'Rule Value':
            for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx):
                for row in cell:
                    row.fill = light_blue_fill

        elif header == 'SERFF Copay Value':
            for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx):
                for row in cell:
                    row.fill = light_orange_fill

        elif header == 'SERFF Deductible Value':
            for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx):
                for row in cell:
                    row.fill = light_orange_fill

        elif header == 'SERFF MOOP Value':
            for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx):
                for row in cell:
                    row.fill = light_orange_fill


    print(sheet_name + ' has been painted')
    # save workbook
    workbook.save(file_path)

# recolor_sheets('benefits')


print("fixing paint")